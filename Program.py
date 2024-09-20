from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from datetime import datetime

# Функция для создания скриншота с использованием явных ожиданий
def take_screenshot_with_wait(driver, url, file_name):
    print(f"Taking screenshot of {url}")

    # Загружаем страницу
    print(f"Loading URL: {url}")
    start_time = datetime.now()  # Записываем время начала загрузки страницы
    driver.get(url)

    # Явное ожидание загрузки страницы
    wait = WebDriverWait(driver, 10)  # Ожидание до 10 секунд
    wait.until(EC.visibility_of_element_located((By.TAG_NAME, 'body')))  # Ожидание видимости элемента на странице
    time.sleep(1)  # Дополнительная задержка для полной загрузки страницы (вы можете увеличить это значение)

    # Создаем скриншот
    screenshot_path = f"{file_name}.png"
    driver.save_screenshot(screenshot_path)

    # Вычисляем время загрузки страницы
    end_time = datetime.now()
    load_time = end_time - start_time

    return screenshot_path, start_time, load_time

# Функция для обработки ссылок из Excel с ожиданием загрузки страницы и вставкой скриншотов
def process_links_and_insert_screenshots(excel_file):
    print("Processing links and inserting screenshots...")
    wb = load_workbook(excel_file)
    ws = wb.active
    row_number = 2  # начальный номер строки (со второй строки)

    # Запускаем веб-драйвер Chrome
    chrome_driver_path = "E:\chrome Drive\chromedriver-win64\chromedriver.exe"
    service = Service(chrome_driver_path)
    options = webdriver.ChromeOptions()
    options.add_argument('--window-size=1920,1080')  # Установка размера окна браузера
    driver = webdriver.Chrome(service=service, options=options)

    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        url = row[0]

        # Создаем уникальное имя для скриншота
        screenshot_file_name = f"screenshot_{row_number}"

        # Получаем скриншот, время начала загрузки и время загрузки страницы
        screenshot_path, start_time, load_time = take_screenshot_with_wait(driver, url, screenshot_file_name)

        # Вставляем время открытия и дату во второй столбец
        ws[f'B{row_number}'] = start_time.strftime("%Y-%m-%d %H:%M:%S")

        # Вставляем время загрузки страницы в третий столбец
        ws[f'C{row_number}'] = f"Loaded in {load_time.total_seconds()} seconds"

        # Вставляем скриншот в ячейку рядом с URL
        img = Image(screenshot_path)
        img.width = 500  # Устанавливаем ширину изображения (вы можете настроить это значение)
        img.height = 400  # Устанавливаем высоту изображения (вы можете настроить это значение)
        # Рассчитываем координаты ячейки для вставки изображения
        cell = f'D{row_number}'
        ws.add_image(img, cell)

        # Установка высоты строки
        ws.row_dimensions[row_number].height = 200

        row_number += 1  # увеличиваем номер строки

    # Сохраняем изменения в файле Excel
    print(f"Saving Excel file: {excel_file}")
    wb.save(excel_file)

    # Останавливаем веб-драйвер
    print("Quitting webdriver.")
    driver.quit()

    print("Screenshots inserted successfully.")

# Пример использования
excel_file = "E:\\скрины кож\\test.xlsx"
process_links_and_insert_screenshots(excel_file)
